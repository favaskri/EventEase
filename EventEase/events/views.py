from django.shortcuts import render,redirect,get_object_or_404
from .models import Event,Profile
from .forms import EventForm
from django.contrib.auth.decorators import login_required,user_passes_test
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpResponseForbidden
from tickets.models import Ticket
from django.db.models import Count
from django.db.models import Sum, F
import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse









# Create your views here.


def is_admin(user):
    return user.is_staff


def index(request):
    return render(request,'index.html')

def admin_event_list(request):
    event_data=Event.objects.prefetch_related('tickets', 'user__user').all()
    context={
        'event_data':event_data
    }
   
    return render(request,'layout_admin_event_list.html',context)

def admin_ticket_list(request):
    # ticket_data=Event.objects.prefetch_related('ticket_set', 'user__user').all()
    ticket_data = Event.objects.prefetch_related('tickets', 'user__user').annotate(
        total_ticket_collection=Sum(F('tickets__price'))
    ).all()
    context={
        'ticket_data':ticket_data
    }
    print(context)
    return render(request,'layout_admin_ticket_list.html',context)

@login_required
@user_passes_test(is_admin)
def admin_users_list(request):
    # Aggregate ticket counts for each user and event
    ticket_summary = (
        Ticket.objects.values('user__username', 'event__title')
        .annotate(tickets_purchased=Count('id'))
        .order_by('user__username', 'event__title')
    )

    return render(request, 'layout_admin_users_list.html', {'ticket_summary': ticket_summary})

@login_required
@user_passes_test(is_admin)
def create_event(request):
    if request.method == 'POST':
        form = EventForm(request.POST,request.FILES)
        if form.is_valid():
            event=form.save(commit=False)
            profile,created=Profile.objects.get_or_create(user=request.user)
            event.user=profile
            form.save()
           
            return redirect('display_event') 
    else:
        form = EventForm()  # Initialize an empty form on GET request

    # Render the form, with validation errors if POST data was invalid
    return render(request, 'create_event_layout.html', {'form': form})

    



def event_display(request):
    event_requests = Event.objects.all().order_by('date')
    print(event_requests)
    paginator = Paginator(event_requests,4)  # 4 events per page

    page_number = request.GET.get('page')
    try:
        events_page = paginator.get_page(page_number)  # Get the current page's events
    except PageNotAnInteger:
        events_page = paginator.get_page(1)  # If page is not an integer, show the first page
    except EmptyPage:
        events_page = paginator.get_page(paginator.num_pages)  # If out of range, show the last page

    return render(request, 'event_display_layout.html', {'event_requests': events_page})


@login_required
@user_passes_test(is_admin)
def event_list(request):
    # Retrieve all venue requests (events) from the database
    user=request.user
    profile , created =Profile.objects.get_or_create(user=user)
    # event_list = Event.objects.filter(user=profile)
    event_list = Event.objects.all()
    print(event_list)
    return render(request,'event_list_layout.html',{'event_list':event_list})

@user_passes_test(is_admin)
def update_event(request,pk):
    event=get_object_or_404(Event,pk=pk)
    if request.method == 'POST':
        print(event)
        form = EventForm(request.POST,request.FILES,instance=event)
        if form.is_valid():
            form.save()
            return redirect('event_list')
                   
    else:
        form = EventForm(instance=event)  

    return render(request, 'update_event_layout.html', {'form': form})


@user_passes_test(is_admin)
def delete_event(request,pk):
     
    # user=request.user
    # profile  =Profile.objects.get(user=user)
    # event=get_object_or_404(Event,pk=pk,user=profile)
    event=get_object_or_404(Event,pk=pk)
    
    
    if request.method == 'POST':
        event.delete()
        print(event_list)
        return redirect('event_list')
      
    return redirect('event_list')


def download_ticket_report(request):
    # Fetch data
    ticket_data = Event.objects.prefetch_related('tickets').annotate(
        total_ticket_collection=Sum(F('tickets__price'))
    ).all()

    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Ticket Report"

    # Set the header
    headers = ['Sr_NO','Event Name','Ticket Price','Event Capacity', 'Available Tickets' , 'Total Revenue (₹)','Event Location' ]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header
        sheet.cell(row=1, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

    # Populate the data rows
    for row_num, event in enumerate(ticket_data, start=2):
        sheet.cell(row=row_num, column=1).value = row_num-1
        sheet.cell(row=row_num, column=2).value = event.title
        sheet.cell(row=row_num, column=3).value = event.ticket_price
        sheet.cell(row=row_num, column=4).value = event.capacity
        sheet.cell(row=row_num, column=5).value = event.available_tickets  # Total tickets
        sheet.cell(row=row_num, column=6).value = event.total_ticket_collection or 0
        sheet.cell(row=row_num, column=7).value = event.location  # Total collection

    # Set response headers for downloading
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ticket_report.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    return response



    


